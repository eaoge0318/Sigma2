"""
XLSX Ingestor - openpyxl 版
第一列為欄位標題，每 ROWS_PER_CHUNK 行合成一個 chunk，相鄰 chunk 重疊 OVERLAP_ROWS 行
格式：【工作表】\n欄位1: 值1 | 欄位2: 值2\n欄位1: 值3 | ...
不支援 .xls（舊格式），請另存為 .xlsx 或 .csv
"""
import uuid
import logging
from typing import List, Dict, Any

logger = logging.getLogger(__name__)

ROWS_PER_CHUNK = 1   # 每個 chunk 包含幾行資料 (改為1代表每行獨立)
OVERLAP_ROWS   = 0   # 相鄰 chunk 重疊幾行
def ingest_xlsx(raw: bytes, filename: str) -> Dict[str, Any]:
    """
    解析 Excel (.xlsx)，每 ROWS_PER_CHUNK 行合成一個 chunk。
    第一列自動辨識為欄位標題。
    """
    import openpyxl
    import io

    doc_id = uuid.uuid4().hex[:12]
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)

    text_chunks = []
    chunk_index = 0
    total_sheets = len(wb.sheetnames)

    for sheet_num, sheet_name in enumerate(wb.sheetnames, start=1):
        ws = wb[sheet_name]
        headers = []
        data_rows = []   # [(row_num, formatted_str), ...]

        for row_num, row in enumerate(ws.iter_rows(values_only=True), start=1):
            cells = list(row)

            # 第一列當欄位標題
            if row_num == 1:
                headers = [str(c).strip() if c is not None else f"欄{i+1}"
                           for i, c in enumerate(cells)]
                continue

            # 跳過全空行
            if all(c is None or str(c).strip() == "" for c in cells):
                continue

            # 組合成 "欄位: 值 | 欄位: 值" 格式
            parts = []
            for i, c in enumerate(cells):
                if c is None or str(c).strip() == "":
                    continue
                col_name = headers[i] if i < len(headers) else f"欄{i+1}"
                parts.append(f"{col_name}: {str(c).strip()}")

            if parts:
                data_rows.append((row_num, " | ".join(parts)))

        if not data_rows:
            continue

        # 滑動視窗：每次取 ROWS_PER_CHUNK 行，步進 ROWS_PER_CHUNK - OVERLAP_ROWS
        step = max(ROWS_PER_CHUNK - OVERLAP_ROWS, 1)
        for start in range(0, len(data_rows), step):
            group = data_rows[start:start + ROWS_PER_CHUNK]
            if not group:
                continue

            first_row = group[0][0]
            last_row  = group[-1][0]
            lines = "\n".join(row_str for _, row_str in group)
            if first_row == last_row:
                content = f"【{sheet_name}】（第 {first_row} 行）\n{lines}"
            else:
                content = f"【{sheet_name}】（第 {first_row}~{last_row} 行）\n{lines}"

            text_chunks.append({
                "id": f"{doc_id}_s{sheet_num}_c{chunk_index}",
                "doc_type": "text_chunk",
                "content": content,
                "metadata": {
                    "source_name": filename,
                    "page": sheet_num,
                    "chunk_index": chunk_index,
                    "file_type": "xlsx",
                    "sheet": sheet_name,
                    "row_start": first_row,
                    "row_end": last_row,
                    "doc_id": doc_id,
                },
            })
            chunk_index += 1

    wb.close()
    logger.info(f"[XLSX] '{filename}' 完成：sheets={total_sheets}，chunks={len(text_chunks)}（每{ROWS_PER_CHUNK}行一chunk，重疊{OVERLAP_ROWS}行）")

    return {
        "doc_id": doc_id,
        "filename": filename,
        "text_chunks": text_chunks,
        "image_chunks": [],
        "manifest": {
            "total_pages": total_sheets,
            "total_text_chunks": len(text_chunks),
            "total_images": 0,
        },
    }
